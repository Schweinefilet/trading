import unittest
import shutil
import time
import os
from pathlib import Path
import openpyxl
import pandas as pd
from datetime import datetime

import config
import webull_watcher

# Define test paths
TEST_ROOT = Path("e:/trading/webull_tracker/test_env")
TEST_DOWNLOADS = TEST_ROOT / "Downloads"
TEST_OUTPUT = TEST_ROOT / "Documents" / "test_log.xlsx"

class TestWebullTracker(unittest.TestCase):
    def setUp(self):
        # Clean up
        if TEST_ROOT.exists():
            shutil.rmtree(TEST_ROOT)
        
        TEST_DOWNLOADS.mkdir(parents=True)
        TEST_OUTPUT.parent.mkdir(parents=True)
        
        # Override config
        config.DOWNLOADS_FOLDER = TEST_DOWNLOADS
        config.MASTER_EXCEL_PATH = TEST_OUTPUT

    def tearDown(self):
        if TEST_ROOT.exists():
           try: 
               shutil.rmtree(TEST_ROOT)
           except:
               pass

    def create_dummy_csv(self, filename, trades):
        filepath = TEST_DOWNLOADS / filename
        df = pd.DataFrame(trades)
        df.to_csv(filepath, index=False)
        return filepath

    def test_pnl_calculation_shorts(self):
        print("\nTesting Short P&L Calculation...")
        
        excel_mgr = webull_watcher.ExcelManager(config.MASTER_EXCEL_PATH)
        processor = webull_watcher.CSVProcessor()
        
        # Scenario: Short
        # Sell 10 @ 150 (Open Short)
        # Buy 10 @ 130 (Close Short)
        # Expected P&L: (150 - 130) * 10 = 200
        
        trades = [
            {'Symbol': 'SHORT', 'Side': 'Sell', 'Quantity': 10, 'Avg Price': 150.0, 'Filled Time': '2023-01-01 10:00:00', 'Status': 'Filled'},
            {'Symbol': 'SHORT', 'Side': 'Buy', 'Quantity': 10, 'Avg Price': 130.0, 'Filled Time': '2023-01-01 11:00:00', 'Status': 'Filled'}
        ]
        path = self.create_dummy_csv("Shorts.csv", trades)
        excel_mgr.append_trades(processor.process_file(path))
        
        wb = openpyxl.load_workbook(config.MASTER_EXCEL_PATH)
        ws = wb["Trade Log"]
        headers = [c.value for c in ws[1]]
        pnl_idx = headers.index("Calculated P/L") + 1
        
        # Row 2: Sell (Open Short, P/L 0)
        # Row 3: Buy (Close Short, P/L 200)
        row3_pnl = ws.cell(row=3, column=pnl_idx).value
        print(f"Calculated P&L for Short Buy-back: {row3_pnl}")
        self.assertAlmostEqual(float(row3_pnl), 200.0, delta=0.01)

    def test_pnl_calculation_long(self):
        print("\nTesting Long P&L Calculation...")
        excel_mgr = webull_watcher.ExcelManager(config.MASTER_EXCEL_PATH)
        processor = webull_watcher.CSVProcessor()
        
        # Buy 10 @ 100, Buy 10 @ 110, Sell 15 @ 120 -> P&L 250
        trades = [
            {'Symbol': 'LONG', 'Side': 'Buy', 'Quantity': 10, 'Price': 100.0, 'Filled Time': '2023-01-01 10:00:00', 'Status': 'Filled'},
            {'Symbol': 'LONG', 'Side': 'Buy', 'Quantity': 10, 'Price': 110.0, 'Filled Time': '2023-01-02 10:00:00', 'Status': 'Filled'},
            {'Symbol': 'LONG', 'Side': 'Sell', 'Quantity': 15, 'Price': 120.0, 'Filled Time': '2023-01-03 10:00:00', 'Status': 'Filled'}
        ]
        path = self.create_dummy_csv("Longs.csv", trades)
        excel_mgr.append_trades(processor.process_file(path))
        
        wb = openpyxl.load_workbook(config.MASTER_EXCEL_PATH)
        ws = wb["Trade Log"]
        headers = [c.value for c in ws[1]]
        pnl_idx = headers.index("Calculated P/L") + 1
        row4_pnl = ws.cell(row=4, column=pnl_idx).value
        self.assertAlmostEqual(float(row4_pnl), 250.0, delta=0.01)

    def test_pnl_special_chars(self):
        print("\nTesting P&L with '@' and 'Filled' column...")
        excel_mgr = webull_watcher.ExcelManager(config.MASTER_EXCEL_PATH)
        processor = webull_watcher.CSVProcessor()
        
        # Test CSV structure similar to user's file
        trades = [
            {'Symbol': 'SPECIAL', 'Side': 'Buy', 'Filled': '10', 'Price': '@100.0', 'Filled Time': '2023-01-01 10:00:00', 'Status': 'Filled'},
            {'Symbol': 'SPECIAL', 'Side': 'Sell', 'Filled': '5', 'Price': '120.0', 'Filled Time': '2023-01-01 11:00:00', 'Status': 'Filled'}
        ]
        path = self.create_dummy_csv("Special.csv", trades)
        excel_mgr.append_trades(processor.process_file(path))
        
        wb = openpyxl.load_workbook(config.MASTER_EXCEL_PATH)
        ws = wb["Trade Log"]
        headers = [c.value for c in ws[1]]
        pnl_idx = headers.index("Calculated P/L") + 1
        qty_idx = headers.index("Qty") + 1
        
        # Check Qty was parsed from 'Filled'
        row2_qty = ws.cell(row=2, column=qty_idx).value
        self.assertEqual(float(row2_qty), 10.0)
        
        # Check P&L (Buy 10@100, Sell 5@120 -> P&L 100 on 5 units)
        row3_pnl = ws.cell(row=3, column=pnl_idx).value
        self.assertAlmostEqual(float(row3_pnl), 100.0, delta=0.01)

    def test_daily_performance_sheet(self):
        print("\nTesting Daily Performance Sheet (Weekdays & Layout)...")
        excel_mgr = webull_watcher.ExcelManager(config.MASTER_EXCEL_PATH)
        processor = webull_watcher.CSVProcessor()
        
        # Two trades on different weekdays (Fri and Mon)
        trades = [
            {'Symbol': 'A', 'Side': 'Buy', 'Quantity': 1, 'Price': 10, 'Filled Time': '2023-01-06 10:00', 'Status': 'Filled'}, # Friday
            {'Symbol': 'A', 'Side': 'Sell', 'Quantity': 1, 'Price': 20, 'Filled Time': '2023-01-09 10:00', 'Status': 'Filled'} # Monday
        ]
        path = self.create_dummy_csv("Test.csv", trades)
        excel_mgr.append_trades(processor.process_file(path))
        
        wb = openpyxl.load_workbook(config.MASTER_EXCEL_PATH)
        ws = wb["Daily Performance"]
        
        # Check headers
        headers = [c.value for c in ws[1]]
        self.assertIn("Total P&L (USD)", headers)
        self.assertNotIn("Money In (USD)", headers)
        self.assertNotIn("Money In Market (VND)", headers)
        
        # Check weekday expansion
        # Fri (6th), Sat(skip), Sun(skip), Mon(9th) -> Should have Fri and Mon (and any in between if we used freq='B')
        # freq='B' is business days.
        # Jan 6 2023 is Friday.
        # Jan 7 is Sat.
        # Jan 8 is Sun.
        # Jan 9 is Mon.
        # So it should have 2 rows if start=6th, end=9th freq='B'? 
        # Actually pd.date_range freq='B' includes start and end if they are business days.
        
        row_count = ws.max_row - 1
        print(f"Daily Performance Rows: {row_count}")
        # Expected rows: Jan 6, Jan 9.
        self.assertEqual(row_count, 2)
        
        # Check rounding
        # P&L for Sell is 10. Check if rounded (it is in our logic)
        pnl_val = ws.cell(row=3, column=headers.index("Date P&L (USD)")+1).value
        self.assertEqual(pnl_val, 10.0)

if __name__ == '__main__':
    unittest.main()
