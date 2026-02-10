import sys
import time
import logging
import hashlib
import pandas as pd
from pathlib import Path
from datetime import datetime
from collections import deque, defaultdict
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference, Series

import config

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

class PnLCalculator:
    def __init__(self):
        # Stores [ {qty, price, date} ]
        self.long_inventory = defaultdict(deque)
        self.short_inventory = defaultdict(deque)

    def _to_float(self, val):
        if val is None or str(val).strip() == "":
            return 0.0
        try:
            # Remove currency symbols, commas, or @ symbols if any
            clean_val = str(val).replace('$', '').replace(',', '').replace('@', '').strip()
            return float(clean_val)
        except:
            return 0.0

    def calculate(self, trades):
        """
        Iterates through trades (sorted by time) and calculates Realized P&L.
        Handles both Long (Buy-to-Open, Sell-to-Close) and Short (Sell-to-Open, Buy-to-Close).
        """
        # Sort by Time carefully
        try:
            trades.sort(key=lambda x: pd.to_datetime(x['Time']) if x.get('Time') else datetime.min)
        except Exception as e:
            logger.warning(f"Sort failed: {e}")
            
        for trade in trades:
            symbol = trade.get('Symbol')
            if not symbol: continue
            
            side = str(trade.get('Side', '')).lower()
            status = str(trade.get('Status', '')).lower()
            
            # Skip non-filled orders if status is provided
            if status and 'filled' not in status:
                continue

            qty = self._to_float(trade.get('Qty'))
            if qty == 0:
                qty = self._to_float(trade.get('Filled Qty'))
            
            # Use Avg Price primarily, fallback to Price
            price = self._to_float(trade.get('Avg Price'))
            if price == 0:
                price = self._to_float(trade.get('Price'))
            
            trade['Calculated P/L'] = 0.0
            if qty <= 0: continue

            if 'buy' in side:
                # 1. Check if we are closing a SHORT
                if self.short_inventory[symbol]:
                    trade_pnl = 0.0
                    remaining_qty = qty
                    while remaining_qty > 1e-6 and self.short_inventory[symbol]:
                        open_slot = self.short_inventory[symbol][0]
                        matched_qty = min(remaining_qty, open_slot['qty'])
                        
                        # Short P&L = (OpenPrice - ClosePrice) * Qty
                        pnl_chunk = (open_slot['price'] - price) * matched_qty
                        trade_pnl += pnl_chunk
                        
                        remaining_qty -= matched_qty
                        open_slot['qty'] -= matched_qty
                        if open_slot['qty'] <= 1e-6:
                            self.short_inventory[symbol].popleft()
                    
                    trade['Calculated P/L'] = trade_pnl
                    # If we bought more than the short, the rest opens a LONG
                    if remaining_qty > 1e-6:
                        self.long_inventory[symbol].append({'qty': remaining_qty, 'price': price, 'date': trade['Time']})
                else:
                    # 2. Open LONG
                    self.long_inventory[symbol].append({'qty': qty, 'price': price, 'date': trade['Time']})
                    
            elif 'sell' in side:
                # 1. Check if we are closing a LONG
                if self.long_inventory[symbol]:
                    trade_pnl = 0.0
                    remaining_qty = qty
                    while remaining_qty > 1e-6 and self.long_inventory[symbol]:
                        open_slot = self.long_inventory[symbol][0]
                        matched_qty = min(remaining_qty, open_slot['qty'])
                        
                        # Long P&L = (ClosePrice - OpenPrice) * Qty
                        pnl_chunk = (price - open_slot['price']) * matched_qty
                        trade_pnl += pnl_chunk
                        
                        remaining_qty -= matched_qty
                        open_slot['qty'] -= matched_qty
                        if open_slot['qty'] <= 1e-6:
                            self.long_inventory[symbol].popleft()
                    
                    trade['Calculated P/L'] = trade_pnl
                    # If we sold more than the long, the rest opens a SHORT
                    if remaining_qty > 1e-6:
                        self.short_inventory[symbol].append({'qty': remaining_qty, 'price': price, 'date': trade['Time']})
                else:
                    # 2. Open SHORT
                    self.short_inventory[symbol].append({'qty': qty, 'price': price, 'date': trade['Time']})
                    
        return trades

class ExcelManager:
    def __init__(self, filepath):
        self.filepath = Path(filepath)
        self.processed_hashes = set()
        self.wb = None
        self._load_or_create()

    def _load_or_create(self):
        if self.filepath.exists():
            try:
                self.wb = openpyxl.load_workbook(self.filepath)
                self._load_hashes()
                logger.info(f"Loaded existing Excel: {self.filepath}")
            except Exception as e:
                logger.error(f"Failed to load existing Excel: {e}")
                raise
        else:
            self.wb = openpyxl.Workbook()
            self._init_sheets()
            self.wb.save(self.filepath)
            logger.info(f"Created new Excel: {self.filepath}")

    def _init_sheets(self):
        default = self.wb.active
        self.wb.remove(default)

        # Create Trade Log
        ws_log = self.wb.create_sheet("Trade Log")
        # Standardized headers for the Excel file (restricted as requested)
        self.standard_headers = ["Symbol", "Side", "Qty", "Price", "Avg Price", "Time", "Calculated P/L"]
        ws_log.append(self.standard_headers)
        self._style_header(ws_log)

        # Create Daily Performance
        self.wb.create_sheet("Daily Performance")

        # Create Summary
        self.wb.create_sheet("Summary")
        self._update_summary_sheet()

    def _style_header(self, ws):
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

    def _load_hashes(self):
        if "Trade Log" not in self.wb.sheetnames: return
        ws = self.wb["Trade Log"]
        headers = [c.value for c in ws[1]]
        try:
            hash_idx = headers.index("Hash") + 1
        except ValueError:
            return 
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[hash_idx-1]:
                self.processed_hashes.add(row[hash_idx-1])

    def append_trades(self, trades):
        # Always read full history to ensure FIFO context is complete
        history = self._read_all_trades()
        
        # Identify new trades (dedupe)
        new_trades = [t for t in trades if t['Hash'] not in self.processed_hashes]
        
        # Even if no new trades, we might want to recalc if history is missing P/L?
        # But for efficiency, usually we only trigger on new file.
        # However, to fix user's issue, let's allow recalc even if only updates needed?
        # For now, stick to "if new_trades/files detected".
        
        if not new_trades and not history:
            return 0
            
        full_list = history + new_trades
        
        # Calculate P&L on the *entire* sequence
        calculator = PnLCalculator()
        calculator.calculate(full_list)
        
        # Write EVERYTHING back to sheet
        ws = self.wb["Trade Log"]
        
        # Ensure we use standard headers
        if not hasattr(self, 'standard_headers'):
            self.standard_headers = ["Symbol", "Side", "Qty", "Price", "Avg Price", "Time", "Calculated P/L"]
            
        # Clear and write headers
        ws.delete_rows(1, ws.max_row)
        ws.append(self.standard_headers)
        self._style_header(ws)
            
        # Write all
        for trade in full_list:
            row = []
            for h in self.standard_headers:
                val = trade.get(h, "")
                # Round numeric values to 2 decimal places
                if isinstance(val, (int, float)):
                    val = round(val, 2)
                row.append(val)
            ws.append(row)
            
            # Re-apply styles
            if trade.get('Side'):
                self._style_row(ws, ws.max_row, trade.get('Side'))
            
            if trade.get('Hash'):
                self.processed_hashes.add(trade['Hash'])

        # Update dependent sheets
        if new_trades:
            logger.info(f"✓ Added {len(new_trades)} new records.")
            
        self._update_daily_performance(full_list)
        self._update_summary_sheet()
        self._adjust_widths(ws)
        
        try:
            self.wb.save(self.filepath)
            if new_trades:
                logger.info(f"Saved update to Excel.")
        except PermissionError:
            logger.error("Error: Excel file is open. Cannot save.")
            return 0

        return len(new_trades)

    def _read_all_trades(self):
        if "Trade Log" not in self.wb.sheetnames: return []
        ws = self.wb["Trade Log"]
        headers = [c.value for c in ws[1]]
        trades = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            record = {}
            for idx, val in enumerate(row):
                if idx < len(headers):
                    record[headers[idx]] = val
            trades.append(record)
        return trades

    def _update_daily_performance(self, all_trades_with_pnl=None):
        if "Daily Performance" in self.wb.sheetnames:
            self.wb.remove(self.wb["Daily Performance"])
        ws_daily = self.wb.create_sheet("Daily Performance", 1)

        # Use passed trades if available (avoid re-reading)
        if all_trades_with_pnl:
            df = pd.DataFrame(all_trades_with_pnl)
        else:
            # Fallback
            history = self._read_all_trades()
            if not history: return
            df = pd.DataFrame(history)

        if df.empty: return

        # Use Calculated P/L if available, else P/L
        pnl_col = 'Calculated P/L' if 'Calculated P/L' in df.columns else 'P/L'
        
        df['Time'] = pd.to_datetime(df['Time'], errors='coerce')
        df['Date'] = df['Time'].dt.date
        df[pnl_col] = pd.to_numeric(df[pnl_col], errors='coerce').fillna(0)
        
        # Calculate Daily Stats
        daily_pnl = df.groupby('Date')[pnl_col].sum().reset_index()
        daily_pnl['Date'] = pd.to_datetime(daily_pnl['Date'])
        
        # Generate all weekdays between min and max date
        if not daily_pnl.empty:
            min_date = daily_pnl['Date'].min()
            max_date = daily_pnl['Date'].max()
            # If max_date is not today, we might want to extend to today? 
            # For now, let's stick to the range of trades.
            all_weekdays = pd.date_range(start=min_date, end=max_date, freq='B')
            weekday_df = pd.DataFrame({'Date': all_weekdays})
            
            # Merge with actual data
            daily_stats = pd.merge(weekday_df, daily_pnl, on='Date', how='left').fillna(0)
        else:
            daily_stats = daily_pnl

        daily_stats = daily_stats.sort_values('Date')
        
        # Round P/L to 2 decimals
        daily_stats[pnl_col] = daily_stats[pnl_col].round(2)
        
        daily_stats['Total_PL_USD'] = daily_stats[pnl_col].cumsum().round(2)
        daily_stats['Daily_PL_USD'] = daily_stats[pnl_col]
        
        # VND
        rate = getattr(config, 'USD_VND_RATE', 25400)
        daily_stats['Daily_PL_VND'] = (daily_stats['Daily_PL_USD'] * rate).round(0) # VND usually no decimals
        daily_stats['Total_PL_VND'] = (daily_stats['Total_PL_USD'] * rate).round(0)
        
        daily_stats['Total_Money_USD'] = daily_stats['Total_PL_USD'] # Adjusting as user removed Money In
        daily_stats['Total_Money_VND'] = (daily_stats['Total_Money_USD'] * rate).round(0)

        headers = [
            "Date", "Total P&L (USD)", "Date P&L (USD)", "Total Money (USD)",
            "Total P&L (VND)", "Date P&L (VND)", "Total Money (VND)"
        ]
        ws_daily.append(headers)
        self._style_header(ws_daily)

        for _, row in daily_stats.iterrows():
            ws_daily.append([
                row['Date'].date(),
                row['Total_PL_USD'],
                row['Daily_PL_USD'],
                row['Total_Money_USD'],
                row['Total_PL_VND'],
                row['Daily_PL_VND'],
                row['Total_Money_VND']
            ])

        self._adjust_widths(ws_daily)
        
        # Charts
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "Daily P&L (USD)"
        chart1.y_axis.title = 'P&L ($)'
        chart1.x_axis.title = 'Date'
        data = Reference(ws_daily, min_col=3, min_row=1, max_row=ws_daily.max_row, max_col=3)
        cats = Reference(ws_daily, min_col=1, min_row=2, max_row=ws_daily.max_row)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 4
        ws_daily.add_chart(chart1, "K2")

        chart2 = LineChart()
        chart2.title = "Total P&L Growth (USD)"
        chart2.style = 13
        chart2.y_axis.title = 'Total P&L ($)'
        chart2.x_axis.title = 'Date'
        data2 = Reference(ws_daily, min_col=2, min_row=1, max_row=ws_daily.max_row, max_col=2)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats)
        ws_daily.add_chart(chart2, "K18")


    def _style_row(self, ws, row_idx, side):
        side = str(side).lower()
        fill = None
        if 'buy' in side:
            fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        elif 'sell' in side:
            fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col)
            if fill:
                cell.fill = fill
            cell.border = thin_border

    def _adjust_widths(self, ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter 
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 30)

    def _update_summary_sheet(self):
        ws_sum = self.wb["Summary"]
        ws_log = self.wb["Trade Log"]
        
        ws_sum.delete_rows(1, ws_sum.max_row)
        total_trades = max(0, ws_log.max_row - 1)
        buy_count = 0
        sell_count = 0
        symbols = set()
        
        headers_row = [c.value for c in ws_log[1]] if ws_log.max_row >= 1 else []
        try:
            side_idx = headers_row.index("Side")
            sym_idx = headers_row.index("Symbol")
        except ValueError:
            side_idx = -1
            sym_idx = -1

        if side_idx != -1:
            for row in ws_log.iter_rows(min_row=2, values_only=True):
                side = str(row[side_idx]).lower()
                sym = row[sym_idx]
                if sym: symbols.add(sym)
                if 'buy' in side: buy_count += 1
                if 'sell' in side: sell_count += 1
            
        stats = [
            ["Last Updated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Total Trades", total_trades],
            ["Buy Orders", buy_count],
            ["Sell Orders", sell_count],
            ["Unique Symbols", len(symbols)],
            ["Symbols Traded", ", ".join(sorted(list(symbols))[:20])]
        ]
        
        for stat in stats:
            ws_sum.append(stat)
            
        for row in ws_sum.iter_rows():
            for cell in row:
                cell.font = Font(size=12)
                if cell.column == 1:
                    cell.font = Font(bold=True)
                
        ws_sum.column_dimensions['A'].width = 20
        ws_sum.column_dimensions['B'].width = 50

class CSVProcessor:
    def process_file(self, filepath):
        try:
            logger.info(f"Reading {filepath.name}...")
            df = None
            for enc in ['utf-8', 'latin-1', 'cp1252']:
                try:
                    df = pd.read_csv(filepath, encoding=enc)
                    break
                except UnicodeDecodeError:
                    continue
            
            if df is None:
                logger.error(f"Failed to read {filepath}: Encoding error")
                return []

            if df.empty:
                logger.warning(f"File {filepath} is empty")
                return []

            return self._normalize_and_hash(df, filepath.name)

        except Exception as e:
            logger.error(f"Error processing {filepath}: {e}")
            return []

    def _normalize_and_hash(self, df, filename):
        # 1. Normalize Header (strip whitespace and make case-insensitive)
        df.columns = [str(c).strip() for c in df.columns]
        
        # Mapping variations to our target columns
        mapping = {
            'Symbol': ['Symbol', 'Ticker', 'TICKER', 'SYMBOL'],
            'Side': ['Side', 'Action', 'ACTION', 'SIDE'],
            'Qty': ['Filled', 'Qty', 'Quantity', 'Total Qty', 'QUANTITY', 'QTY', 'Filled Qty', 'FilledQty', 'Unified Qty'],
            'Price': ['Price', 'Price($)', 'Filled Price', 'Trade Price', 'PRICE'],
            'Avg Price': ['Avg Price', 'Average Price', 'AvgPrice', 'AveragePrice', 'Avg Price($)'],
            'Time': ['Time', 'Filled Time', 'Create Time', 'FilledTime', 'CreateTime', 'TIME'],
            'P/L': ['P/L', 'Profit/Loss', 'Profit & Loss', 'Realized P/L', 'RealizedP/L', 'REALIZED P/L'],
            'Status': ['Status', 'Order Status', 'STATUS'],
            'Commission': ['Commission', 'Comm', 'Fee', 'Fees', 'COMMISSION']
        }
        
        final_mapping = {}
        for target, variations in mapping.items():
            for v in variations:
                if v in df.columns:
                    final_mapping[v] = target
                    break # Use first found variation
                    
        df = df.rename(columns=final_mapping)
        
        logger.info(f"Found columns in {filename}: {list(df.columns)}")
        if 'Symbol' not in df.columns:
            logger.warning(f"Could not find 'Symbol' column in {filename}. Mapped columns: {final_mapping}")

        processed_data = []
        target_cols = [
            "Symbol", "Side", "Qty", "Price", "Filled Qty", "Avg Price", 
            "Status", "Order Type", "Time", "Total", "P/L", "Commission"
        ]
        
        for _, row in df.iterrows():
            record = {}
            row_vals_for_hash = []
            
            sym = str(row.get('Symbol', '')).strip()
            if not sym or pd.isna(row.get('Symbol')):
                continue

            for col in target_cols:
                val = row.get(col, "")
                if pd.isna(val): val = ""
                record[col] = val
                row_vals_for_hash.append(str(val))
            
            record['Source File'] = filename
            record['Imported At'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            hash_src = "|".join(row_vals_for_hash)
            record['Hash'] = hashlib.md5(hash_src.encode('utf-8')).hexdigest()[:12]
            
            # Initial P/L from CSV
            try:
                pnl_str = str(row.get('P/L', '0')).replace('$', '').replace(',', '').strip()
                record['Calculated P/L'] = float(pnl_str) if pnl_str else 0.0
            except:
                record['Calculated P/L'] = 0.0
            
            processed_data.append(record)
            
        return processed_data

class WebullHandler(FileSystemEventHandler):
    def __init__(self, excel_manager, processor):
        self.excel_manager = excel_manager
        self.processor = processor

    def on_created(self, event):
        if event.is_directory: return
        path = Path(event.src_path)
        if not path.name.lower().endswith('.csv'): return

        import re
        if not any(re.match(p, path.name, re.IGNORECASE) for p in config.WEBULL_PATTERNS):
            return

        logger.info(f"Detected new Webull CSV: {path.name}")
        time.sleep(2) 
        self.process_file_with_retry(path)

    def process_file_with_retry(self, path):
        for i in range(3):
            try:
                trades = self.processor.process_file(path)
                if trades:
                    count = self.excel_manager.append_trades(trades)
                    if count > 0:
                        logger.info(f"✓ Added {count} new records to {config.MASTER_EXCEL_PATH.name}")
                    else:
                        logger.info(f"No new unique records found in {path.name} (Duplicate)")
                return
            except PermissionError:
                logger.warning(f"File locked, retrying in 2s... ({i+1}/3)")
                time.sleep(2)
            except Exception as e:
                logger.error(f"Error handling file {path.name}: {e}")
                return

def main():
    print("==================================================")
    print("  Webull CSV -> Excel Auto-Processor")
    print("==================================================")
    print(f"Watching: {config.DOWNLOADS_FOLDER}")
    print(f"Output: {config.MASTER_EXCEL_PATH}")
    print("")

    if not config.DOWNLOADS_FOLDER.exists():
        logger.error(f"Downloads folder not found at: {config.DOWNLOADS_FOLDER}")
        print("Please Check config.py")
        return

    config.MASTER_EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    excel_manager = ExcelManager(config.MASTER_EXCEL_PATH)
    processor = CSVProcessor()
    
    # 1. Scan existing
    print(f"Scanning \"{config.DOWNLOADS_FOLDER}\" for existing Webull CSVs...")
    found_count = 0
    import re
    files = sorted([f for f in config.DOWNLOADS_FOLDER.iterdir() if f.is_file()], key=lambda x: x.stat().st_mtime)
    
    for f in files:
        if f.suffix.lower() == '.csv':
            if any(re.match(p, f.name, re.IGNORECASE) for p in config.WEBULL_PATTERNS):
                found_count += 1
                logger.info(f"Processing existing: {f.name}")
                trades = processor.process_file(f)
                excel_manager.append_trades(trades)

    if found_count == 0:
        print("No existing Webull CSVs found matching patterns.")
    else:
        print(f"Processed {found_count} existing files.")

    # 2. Start Watcher
    handler = WebullHandler(excel_manager, processor)
    observer = Observer()
    observer.schedule(handler, str(config.DOWNLOADS_FOLDER), recursive=False)
    observer.start()
    
    print("\nWaiting for Webull CSV exports...")
    print("(Press Ctrl+C to stop)")
    
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()

if __name__ == "__main__":
    main()
